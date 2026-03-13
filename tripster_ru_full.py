#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ПОЛНАЯ структура Tripster.RU (experience.tripster.ru) - русский сайт экскурсий
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_tripster_ru():
    wb = Workbook()
    wb.remove(wb.active)
    
    # Стили
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    category_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    category_font = Font(bold=True, color="FFFFFF", size=11)
    item_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ========== ЛИСТ 1: АРХИТЕКТУРА САЙТА ==========
    ws1 = wb.create_sheet("Архитектура Tripster.RU")
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 50
    
    headers = ["УРОВЕНЬ", "ЭЛЕМЕНТ", "ОПИСАНИЕ", "ПРИМЕР URL"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    architecture = [
        ["1. ГЛАВНАЯ", "Домашняя страница", "Поиск по городам, популярные направления, многодневные туры", "https://experience.tripster.ru/"],
        ["", "", "", ""],
        ["2. КОНТИНЕНТЫ/РЕГИОНЫ", "7 географических зон", "Россия и СНГ, Европа, Азия, Америка, Африка, Австралия, Антарктида", "/destinations/"],
        ["", "Россия и СНГ", "Россия, Беларусь, Казахстан, Армения, Грузия и др.", "/destinations/russia"],
        ["", "Европа", "Италия, Франция, Испания, Германия и др.", "/destinations/europe"],
        ["", "Азия", "Турция, ОАЭ, Таиланд, Вьетнам, Китай, Япония и др.", "/destinations/asia"],
        ["", "Америка", "США, Мексика, Бразилия и др.", "/destinations/america"],
        ["", "Африка", "Египет, Марокко, ЮАР и др.", "/destinations/africa"],
        ["", "Австралия", "Австралия, Новая Зеландия", "/destinations/australia"],
        ["", "Антарктида", "Антарктида", "/destinations/antarctica"],
        ["", "", "", ""],
        ["3. ГОРОДА", "919 городов в 116 странах", "Каждый город - отдельная коллекция экскурсий", "/experience/Saint_Petersburg/"],
        ["", "", "", ""],
        ["4. РУБРИКИ ГОРОДА", "30-100 рубрик на город", "Тематические подборки экскурсий", "/experience/Saint_Petersburg/173-neobyichnyie-marshrutyi/"],
        ["", "Специальные рубрики", "Все, Со скидкой, Новые, Лучшие", "special_offers/, novye/"],
        ["", "Тематические", "По объектам, форматам, темам", "ermitazh/, kreml/, obzornyie/"],
        ["", "Географические", "Пригороды, районы, достопримечательности", "petergof/, carskoe-selo/, kronshtadt/"],
        ["", "", "", ""],
        ["5. ФИЛЬТРЫ", "4 основных фильтра", "Даты, формат, транспорт, цена", "Динамические"],
        ["", "Любые даты, X чел.", "Выбор даты и количества человек", "Фильтр"],
        ["", "Формат проведения", "Индивидуальная, групповая, мини-группа", "Фильтр"],
        ["", "Способ передвижения", "Пешком, на авто, на автобусе, на велосипеде и т.д.", "Фильтр"],
        ["", "Цена", "Диапазон цен", "Фильтр"],
        ["", "", "", ""],
        ["6. ЭКСКУРСИЯ", "Страница экскурсии", "Описание, гид, цена, отзывы, бронирование", "/experience/42934/"],
        ["7. ГИД", "Страница гида", "Профиль гида, все его экскурсии, отзывы", "/guide/373946/"],
        ["8. ТУРЫ", "Многодневные туры", "Туры на несколько дней", "/tours/vietnam/"],
    ]
    
    for row, data in enumerate(architecture, 2):
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row, col, value)
            cell.alignment = left_align
            cell.border = border
            if col == 1 and value and "." in value and not "https" in value:
                cell.fill = category_fill
                cell.font = category_font
    
    # ========== ЛИСТ 2: РУБРИКИ САНКТ-ПЕТЕРБУРГА (30+) ==========
    ws2 = wb.create_sheet("СПб Рубрики")
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 60
    
    headers2 = ["РУБРИКА", "КОЛ-ВО", "URL"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    spb_rubrics = [
        ["=== СПЕЦИАЛЬНЫЕ ===", "", ""],
        ["Все", "1994", "/experience/Saint_Petersburg/"],
        ["Со скидкой", "317", "/experience/Saint_Petersburg/special_offers/"],
        ["Необычные маршруты", "552", "/experience/Saint_Petersburg/173-neobyichnyie-marshrutyi/"],
        ["Новые", "44", "/experience/Saint_Petersburg/37732-novye/"],
        ["Лучшие", "1008", "/experience/Saint_Petersburg/28381-top/"],
        ["Экскурсии 2025", "210", "/experience/Saint_Petersburg/144738-ekskursii-2025/"],
        ["", "", ""],
        
        ["=== ПО ОБЪЕКТАМ/ДОСТОПРИМЕЧАТЕЛЬНОСТЯМ ===", "", ""],
        ["Разводные мосты", "21", "/experience/Saint_Petersburg/7341-razvodnye-mosty/"],
        ["Петропавловская крепость", "105", "/experience/Saint_Petersburg/8976-petropavlovskaya-krepost/"],
        ["Эрмитаж", "144", "/experience/Saint_Petersburg/8924-ermitazh/"],
        ["Русский музей", "37", "/experience/Saint_Petersburg/20188-russkij-muzej/"],
        ["Музей Фаберже", "~50", "/experience/Saint_Petersburg/20178-muzej-faberzhe/"],
        ["Исаакиевский собор", "~80", "/experience/Saint_Petersburg/8973-isaakievskij-sobor/"],
        ["Спас на Крови", "~70", "/experience/Saint_Petersburg/20189-spas-na-krovi/"],
        ["Казанский собор", "~60", "/experience/Saint_Petersburg/9014-kazanskij-sobor/"],
        ["Петропавловский собор", "~40", "/experience/Saint_Petersburg/20576-petropavlovskij-sobor/"],
        ["Крейсер Аврора", "~30", "/experience/Saint_Petersburg/8926-krejser-avrora/"],
        ["Медный всадник", "~50", "/experience/Saint_Petersburg/20572-mednyj-vsadnik/"],
        ["Михайловский замок", "~40", "/experience/Saint_Petersburg/8980-mihajlovskij-zamok/"],
        ["Адмиралтейство", "~30", "/experience/Saint_Petersburg/9016-admiraltejstvo/"],
        ["Дом Зингера", "~25", "/experience/Saint_Petersburg/20186-dom-zingera/"],
        ["Юсуповский дворец", "~50", "/experience/Saint_Petersburg/20167-yusupovskij-dvorec/"],
        ["", "", ""],
        
        ["=== ПРИГОРОДЫ И ОКРЕСТНОСТИ ===", "", ""],
        ["Царское Село", "78", "/experience/Saint_Petersburg/7343-carskoe-selo/"],
        ["Петергоф", "66", "/experience/Saint_Petersburg/4771-ekskursii-v-petergof/"],
        ["Кронштадт", "62", "/experience/Saint_Petersburg/4714-kronshtadt/"],
        ["Екатерининский дворец", "~60", "/experience/Saint_Petersburg/8925-ekaterininskij-dvorec/"],
        ["Янтарная комната", "~40", "/experience/Saint_Petersburg/20181-yantarnaya-komnata/"],
        ["Большой дворец Петергоф", "~50", "/experience/Saint_Petersburg/95755-bolshoj-dvorec-petergof/"],
        ["Царскосельский лицей Пушкина", "~30", "/experience/Saint_Petersburg/41441-carskoselskij-licej-pushkina/"],
        ["Павловский дворец", "~35", "/experience/Saint_Petersburg/20192-pavlovskij-dvorec/"],
        ["Рускеала", "~40", "/experience/Saint_Petersburg/20244-ruskeala/"],
        ["Ладожское озеро", "~25", "/experience/Saint_Petersburg/20248-ladozhskoe-ozero/"],
        ["Кижи", "~20", "/experience/Saint_Petersburg/37137-kizhi/"],
        ["Соловки", "~15", "/experience/Saint_Petersburg/125463-colovki/"],
        ["", "", ""],
        
        ["=== ПО ТИПУ/ФОРМАТУ ===", "", ""],
        ["Обзорные", "109", "/experience/Saint_Petersburg/167-obzornyie/"],
        ["Дворцы и особняки", "331", "/experience/Saint_Petersburg/9022-dvorcy-i-osobnyaki/"],
        ["Дворцы и особняки изнутри", "89", "/experience/Saint_Petersburg/124684-dvorcy-iznutri/"],
        ["Музеи и искусство", "356", "/experience/Saint_Petersburg/170-muzei-i-iskusstvo/"],
        ["Билеты в музеи", "15", "/experience/Saint_Petersburg/16615-bilety-v-muzei/"],
        ["Однодневные", "242", "/experience/Saint_Petersburg/3712-v-drugoj-gorod/"],
        ["Городские экскурсии", "772", "/experience/Saint_Petersburg/145961-gorodskie-ekskursii/"],
        ["Авторские", "386", "/experience/Saint_Petersburg/20092-avtorskie/"],
        ["Для детей", "194", "/experience/Saint_Petersburg/177-dlya-detej/"],
        ["Гастрономические", "53", "/experience/Saint_Petersburg/176-gastronomicheskie/"],
        ["По барам", "25", "/experience/Saint_Petersburg/19619-baryi-i-nochnaya-zhizn/"],
        ["Фотосессии", "71", "/experience/Saint_Petersburg/300-fotosessii/"],
        ["Трансферы", "8", "/experience/Saint_Petersburg/8330-transfer/"],
        ["Сбежать из города", "22", "/experience/Saint_Petersburg/129404-sbezhat-iz-goroda/"],
        ["Дворы, парадные и коммуналки", "118", "/experience/Saint_Petersburg/8558-po-dvoram-i-paradnym/"],
        ["Персоны Петербурга", "98", "/experience/Saint_Petersburg/124871-persony-peterburga/"],
        ["", "", ""],
        
        ["=== ПО РАЙОНАМ/ЛОКАЦИЯМ ===", "", ""],
        ["Невский проспект", "~100", "/experience/Saint_Petersburg/8978-nevskij-prospekt/"],
        ["Дворцовая площадь", "~80", "/experience/Saint_Petersburg/39799-dvorcovaya-ploshad/"],
        ["Зимний дворец", "~70", "/experience/Saint_Petersburg/9012-zimnij-dvorec/"],
        ["Стрелка Васильевского острова", "~60", "/experience/Saint_Petersburg/60565-strelka-vasilevskogo-ostrova/"],
        ["Васильевский остров", "~90", "/experience/Saint_Petersburg/8982-vasilevskij-ostrov/"],
        ["Екатерининский парк", "~50", "/experience/Saint_Petersburg/67080-ekaterininskij-park/"],
        ["Канал Грибоедова", "~45", "/experience/Saint_Petersburg/37119-kanal-griboedova/"],
        ["Заячий остров", "~30", "/experience/Saint_Peterson burg/20245-zayachij-ostrov/"],
        ["Площадь Искусств", "~35", "/experience/Saint_Petersburg/140852-ploshad-iskusstv/"],
        ["Летний сад", "~45", "/experience/Saint_Petersburg/8981-letnij-sad/"],
        ["Ростральные колонны", "~40", "/experience/Saint_Petersburg/66130-rostralnye-kolonny/"],
        ["Чижик-Пыжик", "~20", "/experience/Saint_Petersburg/60566-chizhik-pyzhik/"],
        ["Александровская колонна", "~30", "/experience/Saint_Petersburg/66128-aleksandrovskaya-kolonna/"],
        ["Атланты Эрмитажа", "~25", "/experience/Saint_Petersburg/66129-atlanty-ermitazha/"],
        ["Новая Голландия", "~35", "/experience/Saint_Petersburg/20261-nova-gollandiya/"],
        ["Каменный остров", "~30", "/experience/Saint_Petersburg/37142-kamennyj-ostrov/"],
        ["Крестовский остров", "~25", "/experience/Saint_Petersburg/60554-krestovskij-ostrov/"],
        ["Елагин остров", "~20", "/experience/Saint_Petersburg/60553-elagin-ostrov/"],
        ["", "", ""],
        
        ["=== СПЕЦИАЛЬНЫЕ КОЛЛЕКЦИИ ===", "", ""],
        ["Осень в Петербурге", "~150", "/experience/Saint_Petersburg/140924-osen-v-peterburge/"],
    ]
    
    row = 2
    for data in architecture:
        if "===" in data[0]:
            ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws1.cell(row, 1, data[0])
            cell.fill = category_fill
            cell.font = category_font
            cell.alignment = center_align
            cell.border = border
        else:
            for col, value in enumerate(data, 1):
                cell = ws1.cell(row, col, value)
                cell.alignment = left_align
                cell.border = border
                if col == 1 and value and "." in value and not "http" in value:
                    cell.fill = item_fill
                    cell.font = Font(bold=True)
        row += 1
    
    # ========== ЛИСТ 3: РУБРИКИ МОСКВЫ (30+) ==========
    ws3 = wb.create_sheet("Москва Рубрики")
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 60
    
    headers3 = ["РУБРИКА", "КОЛ-ВО", "URL"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    moscow_rubrics = [
        ["=== СПЕЦИАЛЬНЫЕ ===", "", ""],
        ["Все", "1306", "/experience/Moscow/"],
        ["Со скидкой", "86", "/experience/Moscow/special_offers/"],
        ["Необычные маршруты", "440", "/experience/Moscow/152-neobyichnyie-marshrutyi/"],
        ["Новые", "80", "/experience/Moscow/37731-novye/"],
        ["", "", ""],
        
        ["=== ПО ОБЪЕКТАМ ===", "", ""],
        ["Здание МГУ", "24", "/experience/Moscow/20683-zdanie-mgu/"],
        ["Новодевичье кладбище", "12", "/experience/Moscow/20723-novodeviche-kladbishe/"],
        ["Шоколадная фабрика", "7", "/experience/Moscow/20223-shokoladnaya-fabrika/"],
        ["Москва-Сити", "27", "/experience/Moscow/390-moscow-city/"],
        ["Третьяковская галерея", "26", "/experience/Moscow/20759-tretyakovskaya-galereya/"],
        ["Красная площадь", "140", "/experience/Moscow/9011-krasnaya-ploshad/"],
        ["Московский Кремль", "85", "/experience/Moscow/435-kreml/"],
        ["Сталинские высотки", "20", "/experience/Moscow/8947-stalinskie-vysotki/"],
        ["Красный Октябрь", "15", "/experience/Moscow/20225-krasnyj-oktyabr/"],
        ["Парк Патриот", "5", "/experience/Moscow/20231-park-patriot/"],
        ["Метро", "16", "/experience/Moscow/7700-metro/"],
        ["Усадьба Кусково", "9", "/experience/Moscow/8953-usadba-kuskovo/"],
        ["ВДНХ", "45", "/experience/Moscow/8951-vdnh/"],
        ["Большой театр", "11", "/experience/Moscow/8946-bolshoj-teatr/"],
        ["Храм Христа Спасителя", "~60", "/experience/Moscow/8945-hram-hrista-spasitelya/"],
        ["ГУМ", "~40", "/experience/Moscow/20232-gum/"],
        ["Патриаршие пруды", "~45", "/experience/Moscow/20660-patriarshie-prudy/"],
        ["Храм Василия Блаженного", "~70", "/experience/Moscow/8948-hram-vasiliya-blazhennogo/"],
        ["Воробьёвы горы", "~55", "/experience/Moscow/8950-vorobevy-gory/"],
        ["Александровский сад", "~50", "/experience/Moscow/20230-aleksandrovskij-sad/"],
        ["Зарядье", "~65", "/experience/Moscow/8559-zaryadye/"],
        ["Старый Арбат", "~75", "/experience/Moscow/8954-staryj-arbat/"],
        ["Манежная площадь", "~40", "/experience/Moscow/38107-manezhnaya-ploshad/"],
        ["Новодевичий монастырь", "~50", "/experience/Moscow/8944-novodevichij-monastyr/"],
        ["Мавзолей Ленина", "~35", "/experience/Moscow/20664-mavzolej-lenina/"],
        ["Коломенское", "~60", "/experience/Moscow/20584-kolomenskoe/"],
        ["Царицыно", "~55", "/experience/Moscow/8952-caricyno/"],
        ["", "", ""],
        
        ["=== ПО ФОРМАТУ/ТИПУ ===", "", ""],
        ["Обзорные", "50", "/experience/Moscow/147-obzornyie/"],
        ["Мистические", "37", "/experience/Moscow/8912-misticheskie/"],
        ["Монастыри, церкви, храмы", "157", "/experience/Moscow/35217-monastyri-cerkvi-hramy/"],
        ["Мастер-классы", "24", "/experience/Moscow/18502-master-classi/"],
        ["Для детей", "166", "/experience/Moscow/156-dlya-detej/"],
        ["Квесты", "115", "/experience/Moscow/7958-kvesty/"],
        ["Гастрономические", "51", "/experience/Moscow/424-gastronomicheskie/"],
        ["Фотосессии", "42", "/experience/Moscow/2760-fotosessii/"],
        ["За городом", "108", "/experience/Moscow/2758-za-gorodom/"],
        ["Активный отдых", "21", "/experience/Moscow/17824-activniy-otdih/"],
        ["", "", ""],
        
        ["=== ПО ТЕМАМ ===", "", ""],
        ["Булгаков", "31", "/experience/Moscow/8910-bulgakov/"],
        ["Сбежать из центра", "22", "/experience/Moscow/129403-sbezhat-iz-centra/"],
        ["Иваново", "9", "/experience/Moscow/36856-ivanovo/"],
    ]
    
    row = 2
    for data in moscow_rubrics:
        if "===" in data[0]:
            ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell = ws3.cell(row, 1, data[0])
            cell.fill = category_fill
            cell.font = category_font
            cell.alignment = center_align
            cell.border = border
        else:
            for col, value in enumerate(data, 1):
                cell = ws3.cell(row, col, value)
                cell.alignment = left_align if col != 2 else center_align
                cell.border = border
                if col == 1 and value and value not in ["", "Все"]:
                    cell.font = Font(bold=True)
        row += 1
    
    # ========== ЛИСТ 4: ФИЛЬТРЫ ==========
    ws4 = wb.create_sheet("Фильтры")
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 70
    ws4.column_dimensions['C'].width = 40
    
    headers4 = ["ФИЛЬТР", "ОПИСАНИЕ", "ОПЦИИ"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    filters = [
        ["Любые даты, X чел.", "Выбор даты проведения и количества участников", "Календарь + счётчик человек"],
        ["Формат проведения", "Тип экскурсии по количеству участников", "Индивидуальная, Групповая, Мини-группа"],
        ["Способ передвижения", "Как будет проходить экскурсия", "Пешком, На автомобиле, На автобусе, На велосипеде, На лодке/катере, На метро, Смешанный и др."],
        ["Цена", "Диапазон цены экскурсии", "Слайдер от-до"],
        ["Фильтры (дополнительные)", "Кнопка открывает расширенные фильтры", "Рейтинг, язык, длительность и др."],
    ]
    
    for row, data in enumerate(filters, 2):
        for col, value in enumerate(data, 1):
            cell = ws4.cell(row, col, value)
            cell.alignment = left_align
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
                cell.fill = item_fill
    
    # ========== ЛИСТ 5: ПОПУЛЯРНЫЕ ГОРОДА ==========
    ws5 = wb.create_sheet("Топ Города")
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 20
    ws5.column_dimensions['C'].width = 20
    ws5.column_dimensions['D'].width = 50
    
    headers5 = ["ГОРОД", "КОЛ-ВО ЭКСКУРСИЙ", "СТРАНА", "URL"]
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    top_cities = [
        ["Санкт-Петербург", "1996", "Россия", "/experience/Saint_Petersburg/"],
        ["Москва", "1308", "Россия", "/experience/Moscow/"],
        ["Калининград", "626", "Россия", "/experience/Kaliningrad/"],
        ["Стамбул", "527", "Турция", "/experience/Istanbul/"],
        ["Тбилиси", "493", "Грузия", "/experience/Tbilisi/"],
        ["Казань", "407", "Россия", "/experience/Kazan/"],
        ["Минск", "337", "Беларусь", "/experience/Minsk/"],
        ["Мурманск", "274", "Россия", "/experience/Murmansk/"],
        ["Дубай", "241", "ОАЭ", "/experience/Dubai/"],
    ]
    
    for row, data in enumerate(top_cities, 2):
        for col, value in enumerate(data, 1):
            cell = ws5.cell(row, col, value)
            cell.alignment = center_align if col in [2] else left_align
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
    
    # ========== ЛИСТ 6: ТИПЫ ЭКСКУРСИЙ ==========
    ws6 = wb.create_sheet("Типы Экскурсий")
    ws6.column_dimensions['A'].width = 35
    ws6.column_dimensions['B'].width = 75
    
    headers6 = ["ТИП ЭКСКУРСИИ", "ОПИСАНИЕ"]
    for col, header in enumerate(headers6, 1):
        cell = ws6.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    types = [
        ["По формату проведения:", ""],
        ["  Индивидуальная", "Только для вас и вашей группы"],
        ["  Групповая", "Открытая группа (присоединение к другим туристам)"],
        ["  Мини-группа", "Небольшая группа до 10-15 человек"],
        ["", ""],
        ["По способу передвижения:", ""],
        ["  Пешком", "Пешеходная экскурсия"],
        ["  На автомобиле", "На легковом авто (индивидуальная)"],
        ["  На автобусе", "На автобусе/микроавтобусе (групповая)"],
        ["  На велосипеде", "Велоэкскурсия"],
        ["  На лодке/катере", "Водная прогулка"],
        ["  На метро", "С использованием метро"],
        ["  Смешанный", "Комбинация нескольких способов"],
        ["  В помещении", "Внутри музея/дворца"],
        ["", ""],
        ["По типу контента:", ""],
        ["  Обзорные", "Знакомство с городом, основные достопримечательности"],
        ["  Тематические", "По определённой теме (Булгаков, архитектура, история и т.д.)"],
        ["  Авторские", "Уникальные маршруты от гида"],
        ["  Необычные маршруты", "Нестандартные экскурсии"],
        ["  Музейные", "Экскурсии по музеям с гидом"],
        ["  Дворцовые", "Экскурсии по дворцам и особнякам"],
        ["  Гастрономические", "Food tours, дегустации"],
        ["  По барам", "Bar crawl, ночная жизнь"],
        ["  Мистические", "Мистика, легенды, привидения"],
        ["  Для детей", "Семейные экскурсии"],
        ["  Квесты", "Интерактивные квесты по городу"],
        ["  Фотосессии", "С профессиональным фотографом"],
        ["  Мастер-классы", "Обучающие активности"],
        ["  Однодневные", "Поездки в другие города"],
        ["  Трансферы", "Трансфер с экскурсией"],
        ["  Билеты в музеи", "Покупка билетов с гидом"],
    ]
    
    row = 2
    for data in types:
        if "===" in data[0]:
            ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws6.cell(row, 1, data[0])
            cell.fill = category_fill
            cell.font = category_font
            cell.alignment = center_align
            cell.border = border
        else:
            for col, value in enumerate(data, 1):
                cell = ws6.cell(row, col, value)
                cell.alignment = left_align
                cell.border = border
                if value and ":" in value:
                    cell.font = Font(bold=True, size=11)
                    cell.fill = item_fill
        row += 1
    
    # ========== ЛИСТ 7: URL СТРУКТУРА ==========
    ws7 = wb.create_sheet("URL Структура")
    ws7.column_dimensions['A'].width = 35
    ws7.column_dimensions['B'].width = 75
    
    headers7 = ["ТИП URL", "ПРИМЕР"]
    for col, header in enumerate(headers7, 1):
        cell = ws7.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    urls = [
        ["Главная", "https://experience.tripster.ru/"],
        ["Все направления", "https://experience.tripster.ru/destinations/"],
        ["Страна/регион", "https://experience.tripster.ru/destinations/russia"],
        ["Город (все экскурсии)", "https://experience.tripster.ru/experience/Saint_Petersburg/"],
        ["Спецпредложения города", "https://experience.tripster.ru/experience/Saint_Petersburg/special_offers/"],
        ["Рубрика (по ID и slug)", "https://experience.tripster.ru/experience/Saint_Petersburg/173-neobyichnyie-marshrutyi/"],
        ["Рубрика достопримечательности", "https://experience.tripster.ru/experience/Saint_Petersburg/8924-ermitazh/"],
        ["Специальная коллекция", "https://experience.tripster.ru/experience/Saint_Petersburg/140924-osen-v-peterburge/"],
        ["Страница экскурсии", "https://experience.tripster.ru/experience/42934/"],
        ["Страница гида", "https://experience.tripster.ru/guide/373946/"],
        ["Многодневные туры (страна)", "https://experience.tripster.ru/tours/vietnam/"],
        ["Многодневные туры (регион)", "https://experience.tripster.ru/tours/russia/kareliya/"],
        ["Избранное пользователя", "https://experience.tripster.ru/favorites/"],
    ]
    
    for row, data in enumerate(urls, 2):
        for col, value in enumerate(data, 1):
            cell = ws7.cell(row, col, value)
            cell.alignment = left_align
            cell.border = border
            if col == 1:
                cell.font = Font(bold=True)
    
    # ========== ЛИСТ 8: СТАТИСТИКА ==========
    ws8 = wb.create_sheet("Статистика")
    ws8.column_dimensions['A'].width = 40
    ws8.column_dimensions['B'].width = 70
    
    headers8 = ["МЕТРИКА", "ЗНАЧЕНИЕ"]
    for col, header in enumerate(headers8, 1):
        cell = ws8.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    stats = [
        ["ВСЕГО ГОРОДОВ", "919 городов"],
        ["ВСЕГО СТРАН", "116 стран"],
        ["КОНТИНЕНТОВ/РЕГИОНОВ", "7 (Россия/СНГ, Европа, Азия, Америка, Африка, Австралия, Антарктида)"],
        ["", ""],
        ["РУБРИК В СПБ", "~30+ видимых + ещё больше скрытых"],
        ["ЭКСКУРСИЙ В СПБ", "1994 экскурсии"],
        ["РУБРИК В МОСКВЕ", "~30+ видимых"],
        ["ЭКСКУРСИЙ В МОСКВЕ", "1306 экскурсий"],
        ["", ""],
        ["ТИПЫ ФИЛЬТРОВ", "4 основных (даты, формат, транспорт, цена) + расширенные"],
        ["ФОРМАТЫ ПРОВЕДЕНИЯ", "3 (Индивидуальная, Групповая, Мини-группа)"],
        ["СПОСОБЫ ПЕРЕДВИЖЕНИЯ", "8+ (Пешком, На авто, На автобусе, На велосипеде, На лодке, На метро, Смешанный, В помещении)"],
        ["", ""],
        ["СТРУКТУРА РУБРИК", "По объектам + По типу + По теме + Специальные"],
    ]
    
    for row, data in enumerate(stats, 2):
        for col, value in enumerate(data, 1):
            cell = ws8.cell(row, col, value)
            cell.alignment = left_align
            cell.border = border
            if col == 1 and value:
                cell.font = Font(bold=True)
                cell.fill = item_fill
    
    # ========== ЛИСТ 9: КЛЮЧЕВЫЕ ОТЛИЧИЯ ==========
    ws9 = wb.create_sheet("Выводы")
    ws9.column_dimensions['A'].width = 110
    
    cell = ws9.cell(1, 1, "КЛЮЧЕВЫЕ ОСОБЕННОСТИ TRIPSTER.RU")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
    
    conclusions = [
        "",
        "1. TRIPSTER.RU - это платформа экскурсий от местных гидов",
        "   - НЕ продажа билетов в парки (как Tripster.com)",
        "   - А продажа ЭКСКУРСИЙ с живыми гидами",
        "",
        "2. СТРУКТУРА РУБРИК:",
        "   - Каждый город имеет 30-100 ТЕМАТИЧЕСКИХ рубрик",
        "   - Рубрики делятся на 4 типа:",
        "     a) Специальные (Все, Со скидкой, Новые, Лучшие)",
        "     b) По объектам/достопримечательностям (Эрмитаж, Кремль, Петергоф...)",
        "     c) По формату (Обзорные, Мистические, Гастрономические, Для детей...)",
        "     d) По темам (Булгаков, Персоны города, Дворцы изнутри...)",
        "",
        "3. URL-СТРУКТУРА:",
        "   - /experience/ГОРОД/ - все экскурсии города",
        "   - /experience/ГОРОД/ID-slug/ - рубрика",
        "   - /experience/ID/ - страница экскурсии",
        "   - /guide/ID/ - страница гида",
        "",
        "4. ОСОБЕННОСТИ:",
        "   - Каждая рубрика имеет уникальный ID (например, 8924-ermitazh)",
        "   - Счётчик экскурсий в рубрике (например, \"1994\" в СПб)",
        "   - Сезонные коллекции (Осень в Петербурге, В Москве осень)",
        "   - Рейтинги гидов (от 1 до 5 звёзд)",
        "",
        "5. ГЕОГРАФИЯ:",
        "   - 919 городов по всему миру",
        "   - Фокус на Россию и СНГ (СПб - 1996, Москва - 1306 экскурсий)",
        "   - Также: Турция, Грузия, ОАЭ, Европа, Азия",
        "",
        "6. ДОПОЛНИТЕЛЬНО:",
        "   - Многодневные туры (/tours/)",
        "   - Мобильное приложение",
        "   - Система избранного (/favorites/)",
        "   - Чаты с гидами",
        "",
        "7. ГЛАВНОЕ ОТЛИЧИЕ ОТ TRIPSTER.COM:",
        "   - Tripster.COM = билеты в парки США (Disney, Universal)",
        "   - Tripster.RU = экскурсии с гидами по всему миру",
        "   - Разные компании, разные бизнес-модели!",
    ]
    
    for row, text in enumerate(conclusions, 2):
        cell = ws9.cell(row, 1, text)
        cell.alignment = left_align
        cell.border = border
        if text.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.")):
            cell.font = Font(bold=True, size=12)
            cell.fill = category_fill
            cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    # Сохраняем
    filename = "Tripster_RU_Structure.xlsx"
    wb.save(filename)
    print(f"SUCCESS! Tripster.RU structure created: {filename}")
    print(f"Sheets: {len(wb.sheetnames)}")
    return filename

if __name__ == "__main__":
    create_tripster_ru()


