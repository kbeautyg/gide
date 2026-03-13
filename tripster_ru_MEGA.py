#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_tripster_ru_mega():
    wb = Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    category_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    category_font = Font(bold=True, color="FFFFFF", size=11)
    item_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def write_rubrics(ws, city_name, rubrics_data):
        """Helper to write rubrics to worksheet"""
        headers = ["РУБРИКА", "КОЛ-ВО", "ID", "SLUG"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        row = 2
        for data in rubrics_data:
            if len(data) == 1 and "===" in data[0]:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                cell = ws.cell(row, 1, data[0])
                cell.fill = category_fill
                cell.font = category_font
                cell.alignment = center_align
                cell.border = border
            else:
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row, col, value)
                    cell.alignment = left_align if col != 2 else center_align
                    cell.border = border
                    if col == 1 and value and value not in [""]:
                        cell.font = Font(bold=True)
            row += 1
    
    # ========== ЛИСТ 1: САНКТ-ПЕТЕРБУРГ - ВСЕ РУБРИКИ ==========
    ws_spb = wb.create_sheet("СПб ВСЕ Рубрики")
    ws_spb.column_dimensions['A'].width = 50
    ws_spb.column_dimensions['B'].width = 12
    ws_spb.column_dimensions['C'].width = 15
    ws_spb.column_dimensions['D'].width = 50
    
    spb_all = [
        ["=== СПЕЦИАЛЬНЫЕ ==="],
        ["Все", "1994", "-", "all"],
        ["Со скидкой", "317", "special_offers", "special_offers"],
        ["Необычные маршруты", "552", "173", "neobyichnyie-marshrutyi"],
        ["Новые", "44", "37732", "novye"],
        ["Лучшие", "1008", "28381", "top"],
        ["Экскурсии 2025", "210", "144738", "ekskursii-2025"],
        [""],
        
        ["=== МУЗЕИ И КУЛЬТУРА ==="],
        ["Эрмитаж", "144", "8924", "ermitazh"],
        ["Русский музей", "37", "20188", "russkij-muzej"],
        ["Музеи и искусство", "356", "170", "muzei-i-iskusstvo"],
        ["Билеты в музеи", "15", "16615", "bilety-v-muzei"],
        ["Музей Фаберже", "~50", "20178", "muzej-faberzhe"],
        ["Третьяковская галерея", "~30", "20759", "tretyakovskaya-galereya"],
        ["Главный штаб", "~25", "61178", "glavnyj-shtab"],
        ["Музей Эрарта", "~20", "60562", "muzej-erarta"],
        [""],
        
        ["=== ДВОРЦЫ И ОСОБНЯКИ ==="],
        ["Дворцы и особняки", "331", "9022", "dvorcy-i-osobnyaki"],
        ["Дворцы и особняки изнутри", "89", "124684", "dvorcy-iznutri"],
        ["Зимний дворец", "~70", "9012", "zimnij-dvorec"],
        ["Екатерининский дворец", "~60", "8925", "ekaterininskij-dvorec"],
        ["Большой дворец Петергоф", "~50", "95755", "bolshoj-dvorec-petergof"],
        ["Юсуповский дворец", "~50", "20167", "yusupovskij-dvorec"],
        ["Павловский дворец", "~35", "20192", "pavlovskij-dvorec"],
        ["Михайловский замок", "~40", "8980", "mihajlovskij-zamok"],
        ["Мраморный дворец", "~30", "20257", "mramornyj-dvorec"],
        ["Шереметевский дворец", "~25", "20258", "sheremetevskij-dvorec"],
        ["Аничков дворец", "~25", "20246", "anichkov-dvorec"],
        ["Строгановский дворец", "~20", "20250", "stroganovskij-dvorec"],
        ["Елагин дворец", "~20", "37141", "elagin-dvorec"],
        ["Воронцовский дворец", "~20", "39374", "voroncovskij-dvorec"],
        ["Кушелев-Безбородко", "~15", "39376", "osobnyak-rumyanceva"],
        ["Особняк Матильды Кшесинской", "~15", "40001", "osobnyak-matildy-kshesinskoj"],
        ["Особняк Кельха", "~15", "20780", "osobnyak-kelha"],
        ["Дом Бенуа", "~15", "39802", "dom-benua"],
        [""],
        
        ["=== ХРАМЫ И СОБОРЫ ==="],
        ["Петропавловская крепость", "105", "8976", "petropavlovskaya-krepost"],
        ["Петропавловский собор", "~40", "20576", "petropavlovskij-sobor"],
        ["Исаакиевский собор", "~80", "8973", "isaakievskij-sobor"],
        ["Спас на Крови", "~70", "20189", "spas-na-krovi"],
        ["Казанский собор", "~60", "9014", "kazanskij-sobor"],
        ["Морской собор", "~30", "60601", "morskoj-sobor-svyatitelya-nikolaya-chudotvorca"],
        ["Никольский морской собор", "~25", "60563", "nikolskij-morskoj-sobor"],
        ["Смольный собор", "~30", "20169", "smolnyj-sobor"],
        ["Александро-Невская лавра", "~40", "8979", "aleksandro-nevskaya-lavra"],
        ["Петрикирхе", "~15", "39796", "petrikirhe"],
        ["Хоральная синагога", "~15", "38835", "horalnaya-sinagoga"],
        [""],
        
        ["=== ПРИГОРОДЫ ==="],
        ["Царское Село", "78", "7343", "carskoe-selo"],
        ["Петергоф", "66", "4771", "ekskursii-v-petergof"],
        ["Кронштадт", "62", "4714", "kronshtadt"],
        ["Однодневные (в другие города)", "242", "3712", "v-drugoj-gorod"],
        ["Рускеала", "~40", "20244", "ruskeala"],
        ["Кижи", "~20", "37137", "kizhi"],
        ["Ладожское озеро", "~25", "20248", "ladozhskoe-ozero"],
        ["Соловки", "~15", "125463", "colovki"],
        ["Ладожские шхеры", "~15", "20256", "ladozhskie-shhery"],
        ["Тихвинский монастырь", "~15", "20247", "tihvinskij-monastyr"],
        ["Свирский монастырь", "~15", "20187", "svirskij-monastyr"],
        ["Форт Александр", "~15", "37138", "fort-aleksandr"],
        ["Крепость Орешек", "~20", "8928", "krepost-oreshek"],
        [""],
        
        ["=== ПЛОЩАДИ И УЛИЦЫ ==="],
        ["Невский проспект", "~100", "8978", "nevskij-prospekt"],
        ["Дворцовая площадь", "~80", "39799", "dvorcovaya-ploshad"],
        ["Площадь Искусств", "~35", "140852", "ploshad-iskusstv"],
        ["Канал Грибоедова", "~45", "37119", "kanal-griboedova"],
        ["Варварка", "~30", "39865", "varvarka"],
        ["Кузнецкий мост", "~25", "69436", "kuzneckij-most"],
        [""],
        
        ["=== ОСТРОВА И ПАРКИ ==="],
        ["Стрелка Васильевского острова", "~60", "60565", "strelka-vasilevskogo-ostrova"],
        ["Васильевский остров", "~90", "8982", "vasilevskij-ostrov"],
        ["Заячий остров", "~30", "20245", "zayachij-ostrov"],
        ["Каменный остров", "~30", "37142", "kamennyj-ostrov"],
        ["Крестовский остров", "~25", "60554", "krestovskij-ostrov"],
        ["Елагин остров", "~20", "60553", "elagin-ostrov"],
        ["Кановерский остров", "~15", "39803", "kanonerskij-ostrov"],
        ["Летний сад", "~45", "8981", "letnij-sad"],
        ["Екатерининский парк", "~50", "67080", "ekaterininskij-park"],
        ["Парк Александрия", "~30", "95757", "park-aleksandriya"],
        ["Марсово поле", "~30", "113530", "marsovo-pole"],
        [""],
        
        ["=== ПАМЯТНИКИ И СИМВОЛЫ ==="],
        ["Медный всадник", "~50", "20572", "mednyj-vsadnik"],
        ["Адмиралтейство", "~30", "9016", "admiraltejstvo"],
        ["Александровская колонна", "~30", "66128", "aleksandrovskaya-kolonna"],
        ["Ростральные колонны", "~40", "66130", "rostralnye-kolonny"],
        ["Атланты Эрмитажа", "~25", "66129", "atlanty-ermitazha"],
        ["Чижик-Пыжик", "~20", "60566", "chizhik-pyzhik"],
        ["Сфинксы", "~25", "130177", "sfinksy"],
        ["Крейсер Аврора", "~30", "8926", "krejser-avrora"],
        [""],
        
        ["=== СПЕЦИАЛЬНЫЕ МЕСТА ==="],
        ["Дом Зингера", "~25", "20186", "dom-zingera"],
        ["Янтарная комната", "~40", "20181", "yantarnaya-komnata"],
        ["Домик Петра I", "~25", "60558", "domik-petra-i"],
        ["Новая Голландия", "~35", "20261", "nova-gollandiya"],
        ["Севкабель Порт", "~25", "39801", "sevkabel-port"],
        ["Лахта Центр", "~20", "20175", "lahta-centr"],
        ["Витебский вокзал", "~20", "37145", "vitebskij-vokzal"],
        ["Дом Елисеевых", "~20", "60589", "dom-kupcov-eliseevyh"],
        ["Смоленское кладбище", "~15", "38108", "smolenskoe-kladbishe"],
        ["Царскосельский лицей", "~30", "41441", "carskoselskij-licej-pushkina"],
        [""],
        
        ["=== ТЕАТРЫ И МУЗЫКА ==="],
        ["Александринский театр", "~25", "60564", "aleksandrinskij-teatr"],
        ["Мариинский театр", "~35", "20173", "mariinskij-teatr"],
        [""],
        
        ["=== ПО ТИПУ ЭКСКУРСИИ ==="],
        ["Обзорные", "109", "167", "obzornyie"],
        ["Городские экскурсии", "772", "145961", "gorodskie-ekskursii"],
        ["Авторские", "386", "20092", "avtorskie"],
        ["Для детей", "194", "177", "dlya-detej"],
        ["Гастрономические", "53", "176", "gastronomicheskie"],
        ["По барам", "25", "19619", "baryi-i-nochnaya-zhizn"],
        ["Фотосессии", "71", "300", "fotosessii"],
        ["Трансферы", "8", "8330", "transfer"],
        ["Сбежать из города", "22", "129404", "sbezhat-iz-goroda"],
        ["Дворы, парадные и коммуналки", "118", "8558", "po-dvoram-i-paradnym"],
        ["Персоны Петербурга", "98", "124871", "persony-peterburga"],
        ["Разводные мосты", "21", "7341", "razvodnye-mosty"],
        [""],
        
        ["=== СПЕЦКОЛЛЕКЦИИ ==="],
        ["Осень в Петербурге", "~150", "140924", "osen-v-peterburge"],
    ]
    
    write_rubrics(ws_spb, "Санкт-Петербург", spb_all)
    
    # ========== ЛИСТ 2: МОСКВА - ВСЕ РУБРИКИ ==========
    ws_msk = wb.create_sheet("Москва ВСЕ Рубрики")
    ws_msk.column_dimensions['A'].width = 50
    ws_msk.column_dimensions['B'].width = 12
    ws_msk.column_dimensions['C'].width = 15
    ws_msk.column_dimensions['D'].width = 50
    
    msk_all = [
        ["=== СПЕЦИАЛЬНЫЕ ==="],
        ["Все", "1306", "-", "all"],
        ["Со скидкой", "86", "special_offers", "special_offers"],
        ["Необычные маршруты", "440", "152", "neobyichnyie-marshrutyi"],
        ["Новые", "80", "37731", "novye"],
        [""],
        
        ["=== ГЛАВНЫЕ ОБЪЕКТЫ ==="],
        ["Красная площадь", "140", "9011", "krasnaya-ploshad"],
        ["Московский Кремль", "85", "435", "kreml"],
        ["Храм Христа Спасителя", "~60", "8945", "hram-hrista-spasitelya"],
        ["ГУМ", "~40", "20232", "gum"],
        ["Храм Василия Блаженного", "~70", "8948", "hram-vasiliya-blazhennogo"],
        ["Третьяковская галерея", "26", "20759", "tretyakovskaya-galereya"],
        ["Мавзолей Ленина", "~35", "20664", "mavzolej-lenina"],
        ["Спасская башня", "~30", "20663", "spasskaya-bashnya"],
        ["Успенский собор", "~25", "20725", "uspenskij-sobor"],
        ["Казанский собор", "~25", "20758", "kazanskij-sobor"],
        ["Исторический музей", "~30", "20772", "istoricheskij-muzej"],
        ["Пушкинский музей", "~30", "20767", "pushkinskiy-muzej"],
        ["Новая Третьяковка", "~25", "41422", "novaya-tretyakovskaya-galereya"],
        ["Политехнический музей", "~20", "95713", "politehnicheskij-muzej"],
        ["Большой театр", "11", "8946", "bolshoj-teatr"],
        ["Мариинский дворец", "~20", "9013", "mariinskij-dvorec"],
        [""],
        
        ["=== СОВРЕМЕННЫЕ ДОСТОПРИМЕЧАТЕЛЬНОСТИ ==="],
        ["Москва-Сити", "27", "390", "moscow-city"],
        ["ВДНХ", "45", "8951", "vdnh"],
        ["Зарядье", "~65", "8559", "zaryadye"],
        ["Красный Октябрь", "15", "20225", "krasnyj-oktyabr"],
        ["Башня Федерация", "~20", "20770", "bashnya-federaciya"],
        ["ГЭС-2", "~20", "36152", "ges-2"],
        ["Останкинская башня", "~25", "8560", "ostankinskaya_bashnya"],
        ["Лахта Центр", "~15", "20175", "lahta-centr"],
        ["Зенит Арена", "~15", "20590", "zenit-arena"],
        [""],
        
        ["=== РАЙОНЫ И УЛИЦЫ ==="],
        ["Патриаршие пруды", "~45", "20660", "patriarshie-prudy"],
        ["Старый Арбат", "~75", "8954", "staryj-arbat"],
        ["Новый Арбат", "~30", "38837", "novyj-arbat"],
        ["Манежная площадь", "~40", "38107", "manezhnaya-ploshad"],
        ["Варварка", "~30", "39865", "varvarka"],
        ["Москва-река", "~50", "8556", "Moskva_reka"],
        ["Кузнецкий мост", "~25", "69436", "kuzneckij-most"],
        ["Ивановская горка", "~25", "36812", "ivanovskaya-gorka"],
        [""],
        
        ["=== УСАДЬБЫ И ПАРКИ ==="],
        ["Царицыно", "~55", "8952", "caricyno"],
        ["Коломенское", "~60", "20584", "kolomenskoe"],
        ["Кол omenский дворец", "~30", "36829", "kolomenskij-dvorec"],
        ["Усадьба Кусково", "9", "8953", "usadba-kuskovo"],
        ["Воробьёвы горы", "~55", "8950", "vorobevy-gory"],
        ["Смотровая площадка Воробьёвы горы", "~40", "20694", "smotrovaya-ploshadka-na-vorobyovyh-gorah"],
        ["Парк Горького", "~40", "8956", "park-gorkogo"],
        ["Нескучный сад", "~20", "41021", "neskuchnyj-sad"],
        ["Музеон", "~25", "20730", "muzeon"],
        ["Александровский сад", "~50", "20230", "aleksandrovskij-sad"],
        [""],
        
        ["=== МОНАСТЫРИ ==="],
        ["Новодевичий монастырь", "~50", "8944", "novodevichij-monastyr"],
        ["Новодевичье кладбище", "12", "20723", "novodeviche-kladbishe"],
        ["Донской монастырь", "~20", "20715", "donskoj-monastyr"],
        ["Монастыри, церкви, храмы", "157", "35217", "monastyri-cerkvi-hramy"],
        [""],
        
        ["=== ОСОБНЯКИ И ДОМА ==="],
        ["Дом на набережной", "~30", "36153", "dom-na-naberezhnoj"],
        ["Особняк Рябушинского", "~25", "20682", "osobnyak-ryabushinskogo"],
        ["Дом Пашкова", "~20", "36807", "dom-pashkova"],
        ["Дом Перцовой", "~15", "39838", "dom-percovoj"],
        ["Толстовский дом", "~15", "20262", "tolstovskij-dom"],
        ["Дом учёных", "~15", "39372", "dom-uchenyh"],
        ["Особняк Арсения Морозова", "~15", "20676", "osobnyak-arseniya-morozova"],
        ["Сандуновские бани", "~15", "39839", "sandunovskie-bani"],
        ["Северный речной вокзал", "~15", "39379", "severnyj-rechnoj-vokzal"],
        ["Гостиница Ленинградская", "~10", "36811", "gostinica-leningradskaya"],
        ["Путевой дворец", "~10", "20681", "putevoj-dvorec"],
        ["Гостиница Украина", "~10", "126086", "gostinica-ukraina"],
        ["Метрополь", "~20", "20769", "metropol"],
        ["ЦДУМ", "~15", "20731", "tsum"],
        ["Библиотека имени Ленина", "~15", "20768", "biblioteka-imeni-lenina"],
        [""],
        
        ["=== СПЕЦИАЛЬНЫЕ ДОСТОПРИМЕЧАТЕЛЬНОСТИ ==="],
        ["Сталинские высотки", "20", "8947", "stalinskie-vysotki"],
        ["Здание МГУ", "24", "20683", "zdanie-mgu"],
        ["Поклонная гора", "~30", "8958", "poklonnaya-gora"],
        ["Парк Победы", "~25", "20679", "park-pobedy"],
        ["Музей Гуревича", "~10", "20954", "muzej-gurchenko"],
        ["Океанариум", "~15", "126277", "okeanarium"],
        ["Палаты Романовых", "~20", "8949", "palaty-romanovyh"],
        ["Кремлёвский некрополь", "~20", "-", "-"],
        ["Царь-пушка", "~15", "20724", "tsar-pushka"],
        ["Печатный двор", "~10", "75499", "pechatnyj-dvor"],
        ["Комендантский дом", "~10", "38098", "komendantskij-dom"],
        ["Императорские дворцовые конюшни", "~10", "60599", "imperatorskie-dvorcovye-konyushni"],
        ["Академия Штиглица", "~10", "37126", "akademii-shtiglica"],
        ["Каменноостровский дворец", "~10", "41070", "kamennoostrovskij-dvorec"],
        ["ЗСД (Западный скоростной диаметр)", "~10", "77958", "zapadnyj-skorostnoj-diametr"],
        ["Кутицкое подворье", "~10", "20678", "krutickoe-podvore"],
        ["Академия художеств", "~10", "39789", "academy-of-arts"],
        ["Кресты (тюрьма)", "~10", "20176", "kresty"],
        ["Константиновский дворец", "~10", "20573", "konstantinovskij-dvorec"],
        ["Шуваловский парк", "~10", "40075", "shuvalovskij-park"],
        ["Дворец Меншикова", "~15", "9015", "dvorec-menshikova"],
        ["Таврический дворец", "~10", "20249", "tavricheskij-dvorec"],
        ["Аптека Пеля", "~10", "37134", "apteka-pelya"],
        ["Храм Вооружённых сил", "~10", "41013", "hram-vooruzhennyh-sil"],
        ["Вагань ковское кладбище", "~10", "20766", "vagankovskoe-kladbishe"],
        [""],
        
        ["=== ПО ТИПУ/ФОРМАТУ ==="],
        ["Обзорные", "50", "147", "obzornyie"],
        ["Мистические", "37", "8912", "misticheskie"],
        ["Для детей", "166", "156", "dlya-detej"],
        ["Квесты", "115", "7958", "kvesty"],
        ["Гастрономические", "51", "424", "gastronomicheskie"],
        ["Фотосессии", "42", "2760", "fotosessii"],
        ["Мастер-классы", "24", "18502", "master-classi"],
        ["Активный отдых", "21", "17824", "activniy-otdih"],
        ["За городом", "108", "2758", "za-gorodom"],
        ["Метро", "16", "7700", "metro"],
        [""],
        
        ["=== ПО ТЕМАМ ==="],
        ["Булгаков", "31", "8910", "bulgakov"],
        ["Сбежать из центра", "22", "129403", "sbezhat-iz-centra"],
        ["Парк Патриот", "5", "20231", "park-patriot"],
        ["Шоколадная фабрика", "7", "20223", "shokoladnaya-fabrika"],
        ["Иваново", "9", "36856", "ivanovo"],
        [""],
        
        ["=== СПЕЦКОЛЛЕКЦИИ ==="],
        ["В Москве осень", "~150", "143401", "v-moskve-osen"],
    ]
    
    write_rubrics(ws_msk, "Москва", msk_all)
    
    # ========== ЛИСТ 3: АРХИТЕКТУРА ==========
    ws_arch = wb.create_sheet("Архитектура Tripster.RU")
    ws_arch.column_dimensions['A'].width = 100
    
    arch_cell = ws_arch.cell(1, 1, "АРХИТЕКТУРА TRIPSTER.RU")
    arch_cell.font = header_font
    arch_cell.fill = header_fill
    arch_cell.alignment = center_align
    arch_cell.border = border
    
    arch_text = [
        "",
        "УРОВЕНЬ 1: ГЛАВНАЯ",
        "  - Поиск по городам",
        "  - Популярные направления (Топ-9 городов)",
        "  - Многодневные туры",
        "  - URL: https://experience.tripster.ru/",
        "",
        "УРОВЕНЬ 2: КОНТИНЕНТЫ/РЕГИОНЫ",
        "  - 7 континентов: Россия и СНГ, Европа, Азия, Америка, Африка, Австралия, Антарктида",
        "  - URL: /destinations/",
        "  - URL страны: /destinations/russia",
        "",
        "УРОВЕНЬ 3: ГОРОДА",
        "  - 919 городов в 116 странах",
        "  - URL: /experience/ГОРОД/",
        "  - Примеры: /experience/Saint_Petersburg/, /experience/Moscow/",
        "",
        "УРОВЕНЬ 4: РУБРИКИ",
        "  - 30-100+ рубрик на каждый крупный город",
        "  - URL: /experience/ГОРОД/ID-slug/",
        "  - Примеры: /experience/Moscow/435-kreml/, /experience/Saint_Petersburg/8924-ermitazh/",
        "",
        "ТИПЫ РУБРИК:",
        "  1. СПЕЦИАЛЬНЫЕ:",
        "     - Все экскурсии города",
        "     - Со скидкой",
        "     - Новые",
        "     - Лучшие",
        "     - Необычные маршруты",
        "",
        "  2. ПО ОБЪЕКТАМ/ДОСТОПРИМЕЧАТЕЛЬНОСТЯМ:",
        "     - Эрмитаж, Кремль, Красная площадь",
        "     - Дворцы (Екатерининский, Зимний, Юсуповский...)",
        "     - Храмы (Спас на Крови, Исаакиевский собор...)",
        "     - Музеи (Русский музей, Третьяковка...)",
        "     - Памятники (Медный всадник, Атланты...)",
        "     - Улицы (Невский, Арбат...)",
        "     - Площади (Дворцовая, Красная...)",
        "",
        "  3. ПО ПРИГОРОДАМ:",
        "     - Царское Село, Петергоф, Кронштадт",
        "     - Рускеала, Кижи, Ладожское озеро",
        "     - Владимир, Ярославль, Суздаль",
        "",
        "  4. ПО ФОРМАТУ:",
        "     - Обзорные",
        "     - Городские экскурсии",
        "     - Авторские",
        "     - Для детей",
        "     - Гастрономические",
        "     - По барам",
        "     - Квесты",
        "     - Фотосессии",
        "     - Мастер-классы",
        "     - Мистические",
        "     - Активный отдых",
        "     - Трансферы",
        "",
        "  5. ПО ТЕМАМ:",
        "     - Булгаков, Достоевский, Пушкин",
        "     - Персоны города",
        "     - Дворы и парадные",
        "     - Разводные мосты (для СПб)",
        "",
        "  6. СЕЗОННЫЕ КОЛЛЕКЦИИ:",
        "     - Осень в Петербурге",
        "     - В Москве осень",
        "     - Зима в городе",
        "",
        "УРОВЕНЬ 5: ЭКСКУРСИЯ",
        "  - Страница конкретной экскурсии",
        "  - URL: /experience/ID/",
        "  - Пример: /experience/42934/",
        "",
        "УРОВЕНЬ 6: ГИД",
        "  - Профиль гида, все его экскурсии",
        "  - URL: /guide/ID/",
        "  - Пример: /guide/373946/",
        "",
        "ДОПОЛНИТЕЛЬНО:",
        "  - Многодневные туры: /tours/страна/регион/",
        "  - Избранное: /favorites/",
        "",
        "ФИЛЬТРЫ (НА СТРАНИЦЕ ГОРОДА):",
        "  1. Любые даты, X чел. - дата + количество",
        "  2. Формат проведения - Индивидуальная/Групповая/Мини-группа",
        "  3. Способ передвижения - Пешком/На авто/На автобусе/На велосипеде/На лодке/В помещении",
        "  4. Цена - диапазон цен",
        "  + Дополнительные фильтры (кнопка Фильтры)",
    ]
    
    for row, text in enumerate(arch_text, 2):
        cell = ws_arch.cell(row, 1, text)
        cell.alignment = left_align
        cell.border = border
        if text and not text.startswith("  ") and text != "":
            cell.font = Font(bold=True, size=11)
            cell.fill = item_fill
    
    # ========== ЛИСТ 4: СТАТИСТИКА ==========
    ws_stats = wb.create_sheet("СТАТИСТИКА")
    ws_stats.column_dimensions['A'].width = 45
    ws_stats.column_dimensions['B'].width = 65
    
    stats_headers = ["МЕТРИКА", "ЗНАЧЕНИЕ"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    stats_data = [
        ["ГЕОГРАФИЯ", ""],
        ["Всего городов", "919"],
        ["Всего стран", "116"],
        ["Континентов/регионов", "7"],
        ["", ""],
        ["ТОП ГОРОДА", ""],
        ["Санкт-Петербург", "1994 экскурсии"],
        ["Москва", "1306 экскурсий"],
        ["Калининград", "626 экскурсий"],
        ["Стамбул (Турция)", "527 экскурсий"],
        ["Тбилиси (Грузия)", "493 экскурсии"],
        ["Казань", "407 экскурсий"],
        ["Минск (Беларусь)", "337 экскурсий"],
        ["", ""],
        ["РУБРИКИ", ""],
        ["Рубрик в СПб", "~100+ (30 видимых + много скрытых)"],
        ["Рубрик в Москве", "~80+ (30 видимых + много скрытых)"],
        ["Типов рубрик", "6 (Специальные, По объектам, По пригородам, По формату, По темам, Спецколлекции)"],
        ["", ""],
        ["URL-ПАТТЕРНЫ", ""],
        ["Типов URL", "8 (главная, destinations, город, рубрика, экскурсия, гид, туры, избранное)"],
        ["ID рубрик", "Каждая рубрика имеет уникальный ID (например, 8924-ermitazh)"],
        ["", ""],
        ["ФИЛЬТРЫ", ""],
        ["Основных фильтров", "4 (даты/количество, формат, транспорт, цена)"],
        ["Форматов проведения", "3 (Индивидуальная, Групповая, Мини-группа)"],
        ["Способов передвижения", "8+ (Пешком, На авто, На автобусе, На велосипеде, На лодке, На метро, Смешанный, В помещении)"],
        ["", ""],
        ["КОНТЕНТ", ""],
        ["Рейтинг гидов", "От 1 до 5 звезд (с точностью до сотых)"],
        ["Отзывы", "Показываются на каждой экскурсии"],
        ["Цены", "От 250 руб. до 100,000+ руб."],
    ]
    
    for row, data in enumerate(stats_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws_stats.cell(row, col, value)
            cell.alignment = left_align
            cell.border = border
            if col == 1 and value and value in ["ГЕОГРАФИЯ", "ТОП ГОРОДА", "РУБРИКИ", "URL-ПАТТЕРНЫ", "ФИЛЬТРЫ", "КОНТЕНТ"]:
                cell.font = Font(bold=True, size=12)
                cell.fill = category_fill
            elif col == 1 and value:
                cell.font = Font(bold=True)
                cell.fill = item_fill
    
    # Save
    filename = "Tripster_RU_MEGA.xlsx"
    wb.save(filename)
    print(f"OK! Tripster.RU MEGA created: {filename}")
    print(f"Sheets: {len(wb.sheetnames)}")
    return filename

if __name__ == "__main__":
    create_tripster_ru_mega()


